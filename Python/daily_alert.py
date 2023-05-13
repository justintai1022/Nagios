#!/usr/bin/env python3
# -*- coding: utf-8 -*-
############################################################################################
#    File name: daily_alert.py                                                        #
#    Author: Orsan Hsu & Justin Tai                                                        #
#    Date created: 2020/10/1                                                               #
#    Date last modified: 2023/03/15                                                        #
#    Python Version: 3.8.5                                                                 #
#    Version: 4.4.2                                                                        #
############################################################################################
#    Version 1.0 : Auto Parse  Nagios Hosts & Service alert
#    Version 2.0 : Fix some bug
#    Version 3.0 : Fix some bug and merge Hosts & Service Excel file .
#    Version 4.0 : Merge Unstable alert to Excel file .
#    Version 4.1 : Fix total alert string conversion to int type.
#    Version 4.2 : Fix unstable alarm less than 24 print -1 issue.
#    Version 4.3 : Fix numpy VisibleDeprecationWarning issue .
#    Version 4.3.1 : Ignore VisbleDeprecationWarning message .
#    Version 4.4 : Write to excel if certificate is less than 7 days old  
#    Version 4.4.1 : Adjust the judgment mechanism of certificate validity less than 7 days.
#    Version 4.4.2 : Fix judgement mechanism bug
############################################################################################

import openpyxl 
from datetime import date
import numpy as np
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import time
import os

#設定環境變數
unstable_url = "http://aaa.com.tw/nagiosxi/reports/topalertproducers.php?reportperiod=last24hours&startdate=&enddate=&starttime=1642729046&endtime=1642815446&search=&host=&service=&hostgroup=&servicegroup=&statetype=hard&hostservice=both&state=&export=0&page=1&mode=getpage&records=500"
unstable_data = []
total_alert = 0
#登入199網頁
nagios_password = input("Please Input Nagios Password  \n")
options = webdriver.ChromeOptions()
#Add argument change log to level 3 to aviod selenium bug about USB connect
options.add_argument("–log-level=3")
options.add_argument("--no-sandbox")
options.add_experimental_option('excludeSwitches', ['enable-logging'])
ser=Service('.\chromedriver.exe')
print("Open Chrome ... ")
driver = webdriver.Chrome(service=ser,options=options,service_log_path=os.devnull)
driver.get("http://aaa.com.tw/nagiosxi/login.php")
print("Please Wait 10 s for login Nagios XI")
time.sleep(10)
elem_user = driver.find_element(By.ID,"usernameBox")
elem_user.send_keys("Nagios")
elem_pwd = driver.find_element(By.ID,"passwordBox")
elem_pwd.send_keys(nagios_password)
commit = driver.find_element(By.ID,"loginButton")
commit.click()

print("Please Wait 10 s browse Nagios XI Host Alert")
time.sleep(10)

today = date.today()
d4 = today.strftime("%Y-%m-%d")

##hosts
###############################################################################
driver.get("http://aaa.com.tw/nagiosxi/includes/components/xicore/status.php?sortby=last_state_change&sortorder=asc&host=&hostgroup=&servicegroup=&records=1000&page=1&search=&hostattr=0&serviceattr=0&hoststatustypes=12&servicestatustypes=0&show=hosts")
print("Please Wait 10 s for create Host Excel File")
time.sleep(10)

# 實例化對象excel對象 
excel_obj = openpyxl.Workbook() 

# excel 內當前活躍的sheet工作表 
excel_obj_sheet = excel_obj.active
excel_obj_sheet.title = "Host_Alert_"+d4 

soup = BeautifulSoup(driver.page_source,"html.parser") #將網頁資料以html.parser
tbody3 = soup.select('table.tablesorter tbody tr td')
range_value =  int(len(tbody3)/6)

#print host 數量
print("HOST alarm :  ",range_value)

if range_value != 0:
    tbody3=np.asarray(tbody3,dtype=object)
    list3=tbody3.reshape((range_value,6))
    np.warnings.filterwarnings('ignore', category=np.VisibleDeprecationWarning)
    x=0
    y=0
    for i in range (range_value):
        for j in range (0,6):
            if 'hascomments.png' in str(list3[i][j]) and 'ack.png'in str(list3[i][j]): #by Justin_2023/03/15
                x=x-1
                break
            elif 'hascomments.png' in str(list3[i][j]) and 'downtime.png' in str(list3[i][j]): #by Justin_2023/03/15
                x=x-1
                break
            elif 'hascomments.png' in str(list3[i][j]) and 'nonotifications.png' in str(list3[i][j]):#by Justin_2023/03/15
                x=x-1
                break       
            #print(str(list3[i][j]))
            #print(x)
            if j<1:
                y=j
                excel_obj_sheet.cell(row=x+1, column=y+1, value=list3[i][j].text)
            elif j>=1:
                y=j+1  ##host has no service col , so col + 1 
                if j==2:
                    if 'h' not in list3[i][j].text:
                        excel_obj_sheet.cell(row=x+1, column=1, value="")
                        excel_obj_sheet.cell(row=x+1, column=2, value="")
                        excel_obj_sheet.cell(row=x+1, column=3, value="")
                        x=x-1
                        break
                    else:
                        excel_obj_sheet.cell(row=x+1, column=y+1, value=list3[i][j].text)
                excel_obj_sheet.cell(row=x+1, column=y+1, value=list3[i][j].text)                  
        x+=1
    #文件保存 
    excel_name = 'host&service&unstable_199_%s'%(d4)+'.xlsx'
    excel_obj.save(excel_name)


###############################################################################
##service
###############################################################################
driver.get("http://aaa.com.tw/nagiosxi/includes/components/xicore/status.php?sortby=last_state_change&sortorder=asc&host=&hostgroup=&servicegroup=&records=1000&page=1&search=&hostattr=0&serviceattr=0&hoststatustypes=0&servicestatustypes=28&show=services")
print("Please Wait 10 s to create Service Excel File")
time.sleep(10)

excel_obj = openpyxl.load_workbook('host&service&unstable_199_%s'%(d4)+'.xlsx')

#Create Sheet 2 title : Service_alert_date
excel_obj.create_sheet("Service_alert_"+d4, 1)
# excel 內當前活躍的sheet 2 工作表 
excel_obj.active = 1
excel_obj_sheet = excel_obj.active
##ws2 = excel_obj.create_sheet("Service_alert"+d4, 1)

# 給單元格賦值 
#link = open(file_name2,"r",encoding="utf-8")
soup = BeautifulSoup(driver.page_source,"html.parser")
hostname = soup.select("div.hostname a") #let HTML div tag a element save as sel :<div class="title"></div>
tbody3 = soup.select('table.tablesorter tbody tr td') 

#service has 7 cols
range_value =  int(len(tbody3)/7)

#print how much rows
print("Service alarm :  ",range_value)

#resize one-dimensional list to two-dimensional list
list3=np.reshape(tbody3, (range_value, 7))

x=0
#let null host saved last row data
for i in range (range_value):
    for j in range(0,7):
        if j==0:
            if str(list3[i][j].text) == "" :
                list3[i][j]=list3[i-1][j]
            if 'hascomments.png' in str(list3[i][j]) and 'ack.png'in str(list3[i][j]): #by Justin_2023/03/15
                x=x-1
                break
            elif 'hascomments.png' in str(list3[i][j]) and 'downtime.png' in str(list3[i][j]): #by Justin_2023/03/15
                x=x-1
                break
            elif 'hascomments.png' in str(list3[i][j]) and 'nonotifications.png' in str(list3[i][j]): #by Justin_2023/03/15
                x=x-1
                break
        if j==1:
            if 'hascomments.png' in str(list3[i][j]) and 'ack.png' in str(list3[i][j]): #by Justin_2023/03/15
                x=x-1
                break
            elif 'hascomments.png' in str(list3[i][j]) and 'downtime.png' in str(list3[i][j]): #by Justin_2023/03/15
                x=x-1
                break
            elif 'hascomments.png' in str(list3[i][j]) and 'nonotifications.png' in str(list3[i][j]): #by Justin_2023/03/15
                x=x-1
                break
        if j==3:
            if 'h' not in list3[i][j].text:
                if i==range_value-1:
                    excel_obj_sheet.cell(row=x+1, column=j-2, value="")
                    excel_obj_sheet.cell(row=x+1, column=j-1, value="")
                    excel_obj_sheet.cell(row=x+1, column=j, value="")
                    break
                else:
                    x=x-1
                    break
            #Check the certificate is less than 7 days old and status is warning.
            if 'SSL' in str(list3[i][j-2].text):
                if 'Warning' in (list3[i][j-1].text):
                    if 'd' in str(list3[i][j].text):
                        num=list3[i][j+3].text.split(" ") #2022/10/17 Version 4.4.1 by Justin
                        if int(num[7]) > 7: #2022/10/17 Version 4.4.1 by Justin
                            x=x-1
                            break
                    else:
                        x=x-1
                        break
                        
                

        #print(list3[i][j].text)
        excel_obj_sheet.cell(row=x+1, column=j+1, value=list3[i][j].text)
    x+=1
        
    #文件保存 
    excel_obj.save(excel_name)

###############################################################################
##unstable by Justin Tai
###############################################################################
driver.get(unstable_url)
print("Please Wait 10 s to create Unstable Excel File")
time.sleep(10)

#抓取不穩定告警頁面的第一列的數字
#alert = (int(driver.find_element(By.XPATH,'//*[@id="report"]/div[4]/div[3]/table/tbody/tr[1]/td[1]/a').text))
alert = (int(driver.find_element(By.XPATH,'/html/body/table/tbody/tr[1]/td[1]/a').text))

excel_obj = openpyxl.load_workbook('host&service&unstable_199_%s'%(d4)+'.xlsx')

#Create Sheet 3 title : Unstable_alert_date
excel_obj.create_sheet("Unstable_alert_"+d4, 2)
excel_obj.active = 2
excel_obj_sheet = excel_obj.active

#比對第一列的數字若沒有大於等於24的話，就不動作，若有的話，則開始將事件寫入到excel。
if alert >= 24 :
    soup = BeautifulSoup(driver.page_source,"html.parser")
    tables = soup.find("tbody")
    
    while True:
        trs = tables.find_all("tr")[total_alert]
        total_alert = total_alert+1
        for tr in trs:
            unstable_data.append(tr.getText())
        if int(unstable_data[0]) >= 24: #確認要寫入到excel中抓取到的事件數是否有大於等於24，若沒有則結束程式，若有的話則寫入excel。
            unstable_data[0] = int(unstable_data[0]) #將total alert從字串轉換成整數。2022/01/24 by Justin
            excel_obj_sheet.append (unstable_data)
            del unstable_data[0:4]
            excel_obj.save(excel_name)
        else:
            break

else:
#存入檔案
    excel_obj.save(excel_name)

#print how much rows
if alert < 24: #2022/03/20 V4.2 by_Justin
    print("Unstable alarm :  ",0)
else:
    print("Unstable alarm :  ",total_alert-1)

###########################Program End###########################################
print("Programe End ...")
#################################################################################